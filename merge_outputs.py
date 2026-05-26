import os
import re
import json
import sys
import openpyxl
from openpyxl import load_workbook, Workbook

# Add project root to path
project_dir = os.path.dirname(os.path.abspath(__file__))
if project_dir not in sys.path:
    sys.path.append(project_dir)

import extract_to_excel

# Files to merge as shown in the user's image
files_to_merge = [
    "Part_B1_Technical_Specification_CO2_Urea_1_500.xlsx",
    "Part_B2_Technical_Specification_CO2-Urea_501_1000.xlsx",
    "Part_B3_Technical_Specification_CO2_Urea_1001_2000.xlsx",
    "Part_B4_Technical_Specification_CO2_Urea_2001_3000.xlsx",
    "Part_B5_Technical_Specification_CO2_Urea_3001_3500.xlsx",
    "Part_B10_Technical_Specification_CO2_Urea_6801_7525.xlsx",
    "telecom_extracted_requirements.xlsx"
]

import argparse

def main():
    parser = argparse.ArgumentParser(description="Merge extracted specification files")
    parser.add_argument("--bypass-filtering", action="store_true", help="Bypass all false-positive filtering during merge")
    args = parser.parse_args()

    # Reload config file dynamic changes
    extract_to_excel.reload_config_and_compile()

    print("Starting merging of Excel output files...")

    all_records = []
    seen_keys = set()
    file_counts = {}

    for fname in files_to_merge:
        fpath = os.path.join(project_dir, fname)
        if not os.path.exists(fpath):
            print(f"[WARNING] File not found: {fname}. Skipping.")
            continue
            
        print(f"Reading {fname}...")
        try:
            wb = load_workbook(fpath, data_only=True)
            ws = wb.active
            
            count = 0
            filtered_count = 0
            
            # Row 1 is header, Row 2 is blank, Data starts from Row 3
            for r_idx in range(3, ws.max_row + 1):
                row_vals = [ws.cell(r_idx, c_idx).value for c_idx in range(1, 7)]
                if not any(row_vals):
                    continue
                    
                itb_file = row_vals[0] or ""
                clause_drg = row_vals[1] or ""
                page_num = str(row_vals[2] or "")
                item = row_vals[3] or ""
                requirement = row_vals[4] or ""
                detail = row_vals[5] or ""
                
                # Skip if it is a false positive
                if extract_to_excel.is_false_positive(requirement, bypass=args.bypass_filtering):
                    filtered_count += 1
                    continue
                    
                # Scrub and anonymize
                clean_itb = extract_to_excel.anonymize(str(itb_file).strip())
                clean_clause = extract_to_excel.anonymize(str(clause_drg).strip())
                clean_page = str(page_num).strip()
                clean_requirement = extract_to_excel.anonymize(str(requirement).strip())
                clean_detail = extract_to_excel.anonymize(str(detail).strip())
                
                # Normalize item label using get_item_label
                clean_item = extract_to_excel.get_item_label(clean_requirement)
                
                # Deduplicate by key
                dedup_key = (clean_itb.lower(), clean_clause.lower(), clean_page.lower(), clean_requirement.lower())
                if dedup_key in seen_keys:
                    continue
                    
                seen_keys.add(dedup_key)
                
                record = {
                    "ITB File Name": clean_itb,
                    "Clause or\nDrawing No.": clean_clause,
                    "Page#": clean_page,
                    "Item": clean_item,
                    "Requirement": clean_requirement,
                    "상세 내용": clean_detail
                }
                all_records.append(record)
                count += 1
                
            file_counts[fname] = count
            print(f"  Successfully loaded {count} unique records. Filtered {filtered_count} false positives.")
            
        except Exception as e:
            print(f"[ERROR] Failed to read {fname}: {e}")

    print(f"\nTotal loaded records before sorting: {len(all_records)}")

    # Sort the final list of merged records logically by Category and Page Number
    sorted_records = extract_to_excel.sort_and_arrange_records(all_records)

    # Define output filenames
    merged_xlsx_path = os.path.join(project_dir, "combined_telecom_requirements.xlsx")
    merged_csv_path = os.path.join(project_dir, "combined_telecom_requirements.csv")

    # Save combined Excel spreadsheet
    print(f"\nWriting merged Excel file to {merged_xlsx_path}...")
    try:
        extract_to_excel.generate_excel_table(sorted_records, extract_to_excel.example_xlsx, merged_xlsx_path)
        print("Excel write complete.")
    except Exception as e:
        print(f"[ERROR] Failed to write Excel: {e}")

    # Save combined CSV file
    print(f"Writing merged CSV file to {merged_csv_path}...")
    try:
        import csv
        headers = ["ITB File Name", "Clause or\nDrawing No.", "Page#", "Item", "Requirement", "상세 내용"]
        with open(merged_csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for rec in sorted_records:
                row_data = []
                for h in headers:
                    val = rec[h]
                    if isinstance(val, str):
                        val = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]', '', val)
                    row_data.append(val)
                writer.writerow(row_data)
        print("CSV write complete.")
    except Exception as e:
        print(f"[ERROR] Failed to write CSV: {e}")

    print("\n--- Merging Summary ---")
    for k, v in file_counts.items():
        print(f"{k}: {v} records")
    print(f"Final Merged Unique Records: {len(sorted_records)}")

if __name__ == "__main__":
    main()
