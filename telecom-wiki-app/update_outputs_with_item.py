import os
import json
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define paths
brain_dir = r"C:\Users\jezeh\.gemini\antigravity\brain\f214b7b2-f735-4f5a-84a2-666e18ac890f"
workspace_dir = r"g:\My Drive\Project\ITB Project"

input_json = os.path.join(brain_dir, "telecom_extracted_requirements.json")
output_json = os.path.join(brain_dir, "telecom_extracted_requirements.json")
output_csv_brain = os.path.join(brain_dir, "telecom_extracted_requirements.csv")
output_md_brain = os.path.join(brain_dir, "telecom_data_extraction.md")

output_csv_ws = os.path.join(workspace_dir, "telecom_extracted_requirements.csv")
output_xlsx_ws = os.path.join(workspace_dir, "telecom_extracted_requirements.xlsx")

# Load existing JSON
print("Loading JSON...")
with open(input_json, "r", encoding="utf-8") as f:
    records = json.load(f)

print(f"Loaded {len(records)} records.")

# Update records to insert "Item" as the first key
updated_records = []
for idx, rec in enumerate(records, 1):
    new_rec = {"Item": idx}
    new_rec.update(rec)
    updated_records.append(new_rec)

# Save updated JSON
print("Saving updated JSON...")
with open(output_json, "w", encoding="utf-8") as f:
    json.dump(updated_records, f, indent=2, ensure_ascii=False)

# Save to CSV (brain)
print("Saving CSV to brain...")
headers = ["Item", "Filename", "ITB Document number", "clause or section", "page number", "the exact description verbatim"]
with open(output_csv_brain, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for rec in updated_records:
        writer.writerow([rec[h] for h in headers])

# Save to CSV (workspace)
print("Saving CSV to workspace...")
with open(output_csv_ws, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for rec in updated_records:
        writer.writerow([rec[h] for h in headers])

# Save to Excel (workspace)
print("Saving Excel to workspace...")
wb = Workbook()
ws = wb.active
ws.title = "Extracted Requirements"

# Set grid lines visible
ws.views.sheetView[0].showGridLines = True

# Colors & Fonts
font_family = "Segoe UI"
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Navy Blue
data_font = Font(name=font_family, size=10)
item_font = Font(name=font_family, size=10, bold=True)

thin_side = Side(border_style="thin", color="D3D3D3")
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Write headers
ws.append(headers)

# Apply header formatting
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_all

ws.row_dimensions[1].height = 28

# Write data and format
for row_idx, rec in enumerate(updated_records, 2):
    # Pre-sanitize strings to avoid IllegalCharacterError
    row_data = []
    for h in headers:
        val = rec[h]
        if isinstance(val, str):
            val = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]', '', val)
        row_data.append(val)
        
    ws.append(row_data)
    
    # Format each cell
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font if col_idx != 1 else item_font
        cell.border = border_all
        
        # Alignments
        if col_idx == 1: # Item
            cell.alignment = Alignment(horizontal="center", vertical="top")
        elif col_idx in [2, 3]: # Filename, ITB doc no
            cell.alignment = Alignment(horizontal="left", vertical="top")
        elif col_idx == 4: # Clause/Section
            cell.alignment = Alignment(horizontal="left", vertical="top")
        elif col_idx == 5: # Page number
            cell.alignment = Alignment(horizontal="center", vertical="top")
        else: # Verbatim description
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Auto-fit columns with caps and wrap
max_widths = {
    1: 8,   # Item
    2: 30,  # Filename
    3: 25,  # ITB Document number
    4: 25,  # clause or section
    5: 22,  # page number
    6: 80   # the exact description verbatim
}

for col_idx, width in max_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Enable auto filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(updated_records) + 1}"

# Freeze header row
ws.freeze_panes = "A2"

wb.save(output_xlsx_ws)
print("Excel saved.")

# Save to Markdown table
print("Saving Markdown table...")
md_content = []
md_content.append("# Extracted Telecom Reference Data\n")
md_content.append("This table lists all telecom-related specifications, requirements, and clauses extracted from the available PDF references in the project directory.\n")
md_content.append("## Summary Metrics\n")
md_content.append(f"- **Total Extracted Requirements**: {len(updated_records)} clauses")

# Count by file
file_counts = {}
for rec in updated_records:
    file_counts[rec["Filename"]] = file_counts.get(rec["Filename"], 0) + 1

sorted_files = sorted(file_counts.items(), key=lambda x: x[1], reverse=True)
for fname, count in sorted_files:
    md_content.append(f"  - {fname}: {count} clauses")

md_content.append("\n## Requirement Table\n")

# Build Table headers
md_headers = ["Item", "Filename", "ITB Document number", "Clause or Section", "Page Number", "Exact Description Verbatim"]
md_content.append("| " + " | ".join(md_headers) + " |")
md_content.append("| " + " | ".join([":---" for _ in md_headers]) + " |")

# Build rows
for rec in updated_records:
    # Format description to escape vertical bars and replace newlines/multiple spaces
    desc = rec["the exact description verbatim"]
    desc_clean = desc.replace("|", "&#124;").replace("\n", " ").strip()
    
    # We want filename to be a markdown link to the file if possible
    # e.g., [filename](file:///g:/My Drive/Project/ITB Project/filename)
    filename_url = f"file:///g:/My%20Drive/Project/ITB Project/{rec['Filename'].replace(' ', '%20')}"
    filename_link = f"[{rec['Filename']}]({filename_url})"
    
    row_str = f"| {rec['Item']} | {filename_link} | {rec['ITB Document number']} | {rec['clause or section']} | {rec['page number']} | {desc_clean} |"
    md_content.append(row_str)

with open(output_md_brain, "w", encoding="utf-8") as f:
    f.write("\n".join(md_content) + "\n")
print("Markdown table saved.")
