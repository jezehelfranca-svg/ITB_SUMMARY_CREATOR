import os
import json
import csv
import re
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define paths
brain_dir = r"C:\Users\jezeh\.gemini\antigravity\brain\f214b7b2-f735-4f5a-84a2-666e18ac890f"
workspace_dir = r"g:\My Drive\Project\CTGU"

input_json = os.path.join(brain_dir, "telecom_extracted_requirements.json")
output_json = os.path.join(brain_dir, "telecom_extracted_requirements.json")
output_csv_brain = os.path.join(brain_dir, "telecom_extracted_requirements.csv")
output_md_brain = os.path.join(brain_dir, "telecom_data_extraction.md")
mapping_json = os.path.join(brain_dir, "pdf_page_specs_mapping.json")

output_csv_ws = os.path.join(workspace_dir, "telecom_extracted_requirements.csv")
output_xlsx_ws = os.path.join(workspace_dir, "telecom_extracted_requirements.xlsx")
output_xlsx_brain = os.path.join(brain_dir, "telecom_extracted_requirements.xlsx")
example_xlsx = os.path.join(workspace_dir, "ITB_SUMMARY_EXAMPLE.xlsx")

# Define 40 audited telecom, network, and security-related requirements
new_records = [
    {
        "ITB File Name": "Part_B3_Technical_Specification_CO2_Urea_1001_2000",
        "Clause or Drawing No.": "B773-000-16-50-DS-0024, PLANT COMMUNICATION SYSTEM DATA SHEET",
        "Page#": "108",
        "Item": "PLANT COMMUNICATION SYSTEM DATA SHEE",
        "Requirement": "C. Power Supply Type: UPS",
        "상세 내용": "Backup Time Requirement for Telecom Power Supply, To be confirmed."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "F. TERMINAL POINTS / Table.1 / Location: Simhadri / S.No.11",
        "Page#": "9",
        "Item": "Security/CISF Control Room distance",
        "Requirement": "Distance of CO2 Capture Plant from nearby Security/CISF Control Room '≥3000 meter",
        "상세 내용": "Physical security interface distance. Requirement column is verbatim table text only."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "F. TERMINAL POINTS / Location: Pudimadaka / S.No.7",
        "Page#": "10",
        "Item": "Security/CISF Control Room distance",
        "Requirement": "Distance of Urea Plant B/L from nearby Security/CISF Control Room. '1000 meter",
        "상세 내용": "Physical security interface distance. Requirement column is verbatim table text only."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "A. Technical Specification – Mechanical / 2.g / NB-3",
        "Page#": "27",
        "Item": "Cable trays / Junction Box",
        "Requirement": "Cable trays, Junction Box etc shall be of FRP / GRP material suitable for coastal region,",
        "상세 내용": "Relevant to telecom/security cable containment where applicable."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "A. Technical Specification – Mechanical / 3.a Codes & Standards-2",
        "Page#": "27",
        "Item": "Electrical installations / hazardous areas",
        "Requirement": "All electrical systems and installations shall comply with applicable IEC standards, including IEC 60079 series for hazardous areas, IEC 60364 for electrical installations, and IEC 60529 for ingress protection.",
        "상세 내용": "Relevant to telecom/security electrical installation and equipment protection."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "A. Technical Specification – Mechanical / 3.a Codes & Standards-3",
        "Page#": "27",
        "Item": "Hazardous area / IP / coastal suitability",
        "Requirement": "The hydrogen generation plant shall be provided with electrical equipment and instrumentation suitable for installation in classified hazardous areas (Zone 1 and Zone 2) for Gas Group IIC (Hydrogen service), with appropriate temperature class. Outdoor equipment shall have minimum ingress protection of IP65/IP66, as applicable, and shall be designed with materials, surface treatments, and coatings suitable for the specified ambient and corrosive (coastal) environmental conditions.",
        "상세 내용": "Applies where telecom/security equipment is installed in hydrogen plant hazardous/coastal areas."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "A. Technical Specification – Mechanical / 3.b Design Considerations-5",
        "Page#": "28",
        "Item": "Remote operation / central control room",
        "Requirement": "The electrolyser package and associated hydrogen generation system shall be provided with adequate instrumentation, control, and automation features to enable safe remote operation, continuous monitoring, and supervisory control from the central control room.",
        "상세 내용": "Control/monitoring network and CCR integration relevance."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "A. Technical Specification – Mechanical / 4.b Design Considerations-7",
        "Page#": "30",
        "Item": "Remote monitoring / supervision / control",
        "Requirement": "The Cryogenic ASU shall be provided with adequate instrumentation, control, and automation features to enable safe local operation as well as remote monitoring, supervision, and control of nitrogen and oxygen production from the central control room.",
        "상세 내용": "Control/monitoring network and CCR integration relevance."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "B. Technical Specification – Electrical / 7. ELECTRICAL / 5. Specific requirement for CO2 Capture Block / Item 5",
        "Page#": "37",
        "Item": "DCS/PLC/Instrumentation Redundant UPS",
        "Requirement": "240V, 10kVA AC (Minimum) dual redundant with bypass UPS system (2 nos 1X100%) with Nickel cadmium battery, ACDB, cell booster for DCS/PLC/Instrumentation loads including spare feeders. Battery backup time of 3 hours.",
        "상세 내용": "Dual redundant 10kVA UPS system with Ni-Cd battery bank providing 3 hours backup for DCS/PLC and instrumentation loads at Simhadri CO2 Capture Block."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 1 Hierarchy of Documents & Overview / Clause 3",
        "Page#": "38",
        "Item": "Remote Monitoring & Historian",
        "Requirement": "The bidder shall provide remote monitoring (read-only data flow) for the CO2 capture plant (located at Simhadri) at Pudimadaka. The Pudimadaka standalone monitoring system shall be non-control, with no interface to plant control systems, and will follow cybersecurity-driven network isolation with secure, standardized communication (e.g., OPC) to ensure reliable data monitoring. The system shall include historian with a minimum storage capacity of two months.",
        "상세 내용": "Establish a read-only remote monitoring interface from Simhadri to Pudimadaka using OPC communication, including cybersecurity network isolation and a historian with minimum 2 months storage."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 1 Hierarchy of Documents & Overview / Clause 4",
        "Page#": "38",
        "Item": "Cybersecurity Compliance",
        "Requirement": "The system shall comply with applicable cybersecurity standards like IEC 62443 and include redundant communication using managed switches.",
        "상세 내용": "Mandatory cybersecurity compliance with IEC 62443 standard and network switch redundancy using managed switches."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 3 Design Considerations / Clause 4",
        "Page#": "39",
        "Item": "Remote Connectivity Security",
        "Requirement": "Remote connectivity for the systems, limited to process viewing only. The bidder must supply the necessary hardware and software to enable remote connectivity, ensuring all required cyber security compliance is met as per IEC 62443.",
        "상세 내용": "Remote connections are restricted to process viewing only. Bidder must provide secure hardware/software for remote access compliant with IEC 62443."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 3 Design Considerations / Clause 5",
        "Page#": "39",
        "Item": "System Log Retention",
        "Requirement": "System logs of all DCS/PLC infrastructure involved in the project shall be securely retained for a minimum period of 180 days to support audit and incident investigation requirements.",
        "상세 내용": "Secure storage and retention of system logs for all DCS/PLC network infrastructure for at least 180 days for security auditing."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 3 Design Considerations / Clause 6",
        "Page#": "39-40",
        "Item": "Password Access Control",
        "Requirement": "The system shall have built-in safety features that will allow/disallow certain functions and entry fields within a function to be under password control to protect against inadvertent and unauthorised use of these functions. Assignment of allowable functions and entry fields shall be on the basis of user profile. The system security shall contain various user levels with specific rights which shall be as finalized by the Employer during detailed engineering. However, no. of user levels, no. of users in a level and rights for each level shall be changeable by the programmer (Administrator). The rights of each user shall contain two types of privileges as follows: (a) Privileges for the DCS/PLC, (b) Privileges for the Operating System features. Typically following user levels shall be available: (a) Operator, (b) Supervisor, (c) Maintenance Engineer, (d) Programmer, (e) Shift /Station In charge.",
        "상세 내용": "Role-based access control with password protection across different user levels (Operator, Supervisor, Maintenance, Programmer, Shift In charge) and separate privileges for DCS/PLC and OS."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 3 Design Considerations / Clause 7",
        "Page#": "40",
        "Item": "Control system component redundancy",
        "Requirement": "The redundancy in control system components (Network Switches, Controller, Servers-if any, Power System, Prefab etc.) System shall be designed by the contractor to ensure that malfunction of any single Control system component/ power supply system component etc. shall not lead to loss of any major auxiliary or loss of control function or loss of protection function.",
        "상세 내용": "No single point of failure: redundant network switches, controllers, power supplies, and servers to ensure continuous control and protection."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 3 Design Considerations / Clause 8",
        "Page#": "40",
        "Item": "Conformal Coating",
        "Requirement": "All the electronic modules PCB should have conformal coating that can provide protection against extreme moisture, corrosive gases and aggressive dust, or combinations thereof.",
        "상세 내용": "Conformal coating on all electronic boards/PCBs to protect against coastal corrosion, humidity, and dust."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 3 Design Considerations / Clause 12",
        "Page#": "40",
        "Item": "Network Hardening",
        "Requirement": "Network infrastructure with firewall and cyber security and components shall be designed as per IEC 62443. System hardening must be ensured.",
        "상세 내용": "Network security requirements including firewall implementation, system hardening (disabling unused ports/services), and alignment with IEC 62443."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 3 Design Considerations / Clause 13",
        "Page#": "40",
        "Item": "UPS Monitoring",
        "Requirement": "UPS Systems monitoring signals to be hooked up with centralized DCS/PLC. Qty of signals shall be as recommended by OEM of UPS system.",
        "상세 내용": "Integration of UPS monitoring alarms and status signals with the central DCS/PLC system."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 3 Design Considerations / Clause 19",
        "Page#": "41",
        "Item": "OS & Anti-Virus Updates",
        "Requirement": "The software packages including OS, Application software (eg GUI, History etc ) as per the functional requirement and Anti-Virus Software to be included with the Servers/Workstations/PC Stations (as applicable) shall also be the latest version available at the time of supply. As a customer support, the Contractor shall periodically inform and upgrade the Anti-Virus software of the workstations/servers/switches/firewall as applicable till completion of the warranty period and till the completion of the AMS period.",
        "상세 내용": "Supply latest software and OS versions. Provide periodic anti-virus engine and signature updates for workstations, servers, network switches, and firewalls during warranty and AMS."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 3 Design Considerations / Clause 20",
        "Page#": "41",
        "Item": "History Server Redundancy",
        "Requirement": "History function (Alarm/Event/Trend etc) for the units monitored and controlled from the operator console shall be dual redundant with each node having dual disc drives dedicated for history storage or with history storage available in each of the multiple operator stations of a console group.",
        "상세 내용": "Dual redundant history logging (alarms, events, trends) using dedicated dual disk drives per node or distributed operator stations."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 1",
        "Page#": "41",
        "Item": "DCS/PLC Processor Redundancy",
        "Requirement": "Centralized Controller shall be provided with Dual processor, Hot standby based DCS/PLC system (including main processing unit and memories) one for normal operation and one as hot standby – and should be suitably interfaced with HMI systems.",
        "상세 내용": "DCS/PLC controllers must feature dual processors with hot-standby redundancy for main processing units and memory."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 2",
        "Page#": "41",
        "Item": "HMIPIS Network Interface",
        "Requirement": "Human - Machine Interface & Plant Information System (HMIPIS): Minimum number of HMI PC Three (3)- OWS and One (1) EOWS, and one printer shall be provided for control room along with suitable console and furniture. The profile and dimension shall be decided during detailed engineering and shall be subject to Employer’s approval without any additional cost. PCs shall have Dual Ethernet interface with LAN accessories for all PC based OWS and EWS shall be provided. HMIPIS configured around latest state-of-the art servers/Workstations with open architecture supporting OPC/TCP/IP protocols, etc. shall be provided. This unified HMI can be interfaced either with the control systems of main plant or with propriety/native HMI of systems provided by OEM.",
        "상세 내용": "Provide dual Ethernet network interfaces for all OWS and EWS PCs. System must use open architecture supporting OPC and TCP/IP."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 3",
        "Page#": "41-42",
        "Item": "Large Video Screen Camera Integration",
        "Requirement": "The plant is envisaged to be controlled and monitored from the minimum two Large Video Screens (LVS) in association with its workstation mounted on the Unit Control Desk (UCD), located in the Central Common Control room under all regimes of operation i.e. start-up, shutdown, process control & emergency handling i.e. suitable for plant operation. Large Video Screens (LVS) with its workstation (independent of OWS) and graphic processors to dynamically display plant data / mimics / alarms and any other process information. The Large Video screens shall have additional features to work in association with multiple numbers of plant cameras, if required, also to be supplied under this package.",
        "상세 내용": "Integration of plant surveillance CCTV cameras with the control room Large Video Screens (LVS)."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 4",
        "Page#": "42",
        "Item": "Controller Failover Time",
        "Requirement": "Centralized Controller shall have facility of bump less transfer in case of failure of one. In case of failure of working CPU, standby CPU shall takeover and maximum data loss shall be for 50ms.",
        "상세 내용": "Bumpless failover between redundant controller CPUs with maximum data loss limited to 50 milliseconds."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 5",
        "Page#": "42",
        "Item": "PLC SIL Level Study",
        "Requirement": "SIL level of PLCs (standalone equipment/packages), instruments, solenoid valves etc. shall be decided as per HAZOP and SIL study.",
        "상세 내용": "Safety Integrity Level (SIL) for package PLCs and field instrumentation must be determined through formal HAZOP and SIL studies."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 6",
        "Page#": "42",
        "Item": "UPS Redundancy & Backup",
        "Requirement": "All C&I System including PLC along with OWS (Operator Workstation) & EWS (Engineers Workstation), Analyzers, instruments etc. shall be powered by Redundant UPS (2 X 100 %) along with battery backup in accordance with Electrical specifications. System must be supplied with at least two-hour backup.",
        "상세 내용": "Provide dual redundant 2x100% UPS system with minimum 2-hour battery backup for all C&I systems, PLCs, OWS/EWS, and analyzers."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 7",
        "Page#": "42",
        "Item": "Redundant Communication Bus",
        "Requirement": "Data Communication System Bus connecting Control System and HMIPIS. Other bus systems for connecting various systems/subsystems like Cubicle Bus, Local Bus, I/O Bus (Including Remote I/O Bus) soft links (including those from Field Bus based temperature transmitter) as well as within systems/sub-systems. All the bus systems shall be redundant except for backplane buses which can be non-redundant.",
        "상세 내용": "Redundant system communication buses connecting control systems and HMIPIS, including I/O and cubicle buses."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 8",
        "Page#": "42",
        "Item": "Sequence of Events Recording",
        "Requirement": "Sequence of events recording function for with a resolution of one millisecond with facility of historical storage.",
        "상세 내용": "Sequence of Events Recording (SOER) capability with 1-millisecond resolution for event/alarm analysis."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 9",
        "Page#": "42",
        "Item": "Field Instrument Ingress Protection",
        "Requirement": "All electronic instruments and enclosures in field shall be dust proof and weatherproof to IP-65 as per IEC-60529 or equivalent NEMA enclosure rating or better and secure against the ingress of fumes, dampness, insects, and vermin. All external surfaces shall be suitably treated to provide protection against corrosive plant atmosphere.",
        "상세 내용": "IP65 minimum rating for field instruments, with protection against humidity, dust, vermin, and corrosive coastal atmosphere."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 11",
        "Page#": "42-43",
        "Item": "Coastal Cable Trays & Enclosures",
        "Requirement": "The Control and Instrumentation (C&I) system including conduits / cable sub trays / cabling accessories/ impulse piping / air supply piping / local instrument enclosures and racks/ accessories etc shall be designed in accordance with standards suitable for coastal environments, taking into account factors such as high humidity, salinity, and corrosion potential. All equipment and components must comply with relevant international standards such as IEC, IEEE, and NEMA, with materials and enclosures rated for marine or coastal use (e.g., stainless steel 316, IP66/NEMA 4X). Appropriate protective measures, including conformal coating, sealing, and environmental shielding and the use of corrosion-resistant paints etc. shall be implemented to ensure reliable and long-term operation.",
        "상세 내용": "C&I containment, conduits, racks, and local enclosures must be marine-grade (SS316, IP66, NEMA 4X) to resist coastal salinity and humidity."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 12",
        "Page#": "43",
        "Item": "CCTV Surveillance System",
        "Requirement": "The CCTV System for monitoring of area along with all its components, i.e. necessary hardware and software, interface to the LVS of HMIPIS, the nos. of camera units, servers, network video recorder, network switches, modules and software, any other hardware/ software required for the safe and satisfactory operation, control, protection, monitoring, testing and maintenance of the system shall be provided by the Bidder. The system operation would be covering the complete view of the areas with pan / tilt, zoom, propositioning of the cameras and with programmability to monitor any camera on any monitor either manually or automatically in a defined switching. All cameras shall have CE/FCC/UL/BIS/STQC certification with consideration that cameras shall meet or exceed the statutory requirements as per applicable government norms. The server memory needs to be sufficient for storage of 2 months minimum. The total Camera quantity shall be minimum 20 Nos. Including 12 nos. for outdoor and 08 nos. for indoor. The final number of cameras shall not be limited to the quantities mentioned above. Vendor MUST note that the exact quantity and type (indoor/outdoor) shall be finalized during the detailed engineering, based on site conditions, surveillance coverage requirements without any extra cost to employer.",
        "상세 내용": "IP-based CCTV system with NVR storage of at least 2 months and a minimum camera count of 20 (12 outdoor, 8 indoor)."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 13",
        "Page#": "43",
        "Item": "IP Public Address System (PAGA)",
        "Requirement": "IP Based Public Address system including Call stations, loudspeakers, Network switches, Servers, PA system management Software & other applicable software(s), PC Stations, Interconnecting cable etc. for all the units, common plant areas shall be provided by the Contractor. Minimum Quantity of Master Control Unit (MCU)- 1 nos. (in Centralized Control room), Indoor type calling stations- 3 nos. (with amplifier and loudspeaker), Outdoor type calling station- 5 nos. (with amplifier and loudspeaker), Acoustic hood- 2 nos., All the other items- hardware, software, licenses, including public address system erection hardware, all type of cables, cable tray, junction boxes, racks, conduits, etc. as required for the proper installation (conforming to IS:1881, IS:1882) to make the IP based PA system complete and functional are under Contractor's scope on as required basis.Vendor MUST note that the exact quantity of components shall be finalized during the detailed engineering, based on site conditions , system coverage requirements without any extra cost to employer.",
        "상세 내용": "IP-based PAGA system with 1 Master Control Unit in CCR, 3 indoor and 5 outdoor call stations, and 2 acoustic hoods, meeting IS:1881/1882."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 14",
        "Page#": "43",
        "Item": "GPS Clock Synchronization",
        "Requirement": "Master & Slave Clock System: One Geo-positionary satellite (GPS) based Master Clock in redundant configuration and slave clock for main plant and common areas shall be provided with suitable equipment including antenna, receiver and associated electronics to receive synchronization signals from GPS. The master clock shall synchronize all the systems (main and subsystem/standalone) at suitable intervals to maintain uniform time throughout the CO2 capture plant as per requirement.",
        "상세 내용": "Redundant GPS Master Clock and slave clocks to synchronize time across all control and network systems."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 15",
        "Page#": "43-44",
        "Item": "Separate Power Supply Bus",
        "Requirement": "Separate power supply bus shall be provided for interrogation voltage supply for all inputs and output respectively (even if the input interrogation voltage and output voltage is same).",
        "상세 내용": "Isolate inputs and outputs by providing separate interrogation voltage power supply buses."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 16",
        "Page#": "44",
        "Item": "DCS/PLC Maintenance Training",
        "Requirement": "Training (Minimum 7 days) for troubleshooting, maintenance and modification of centralized DCS/PLC and package PLCs-if any.",
        "상세 내용": "Minimum 7 days of technical training for plant staff covering troubleshooting, maintenance, and modification of the DCS and PLCs."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 17",
        "Page#": "44",
        "Item": "Sub-package communication link",
        "Requirement": "All sub packages with its utilities shall be connected to centralized control room PLC through a two-way communication link.",
        "상세 내용": "Bidder must establish two-way communication links between all package/utility PLCs and the centralized DCS/PLC."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 18",
        "Page#": "44",
        "Item": "Integration signals",
        "Requirement": "For integration purposes the contractor shall determine all the optimal hardwire and soft signals required to achieve data transfer for integration purposes and hence collect this data.",
        "상세 내용": "Identify and implement appropriate hardwired and software-based data interfaces for system integration."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "C. Technical Specification – Control & Instrumentation / 8. CONTROL & INSTRUMENTATION / 4 Control System Architecture / Clause 19",
        "Page#": "44",
        "Item": "Centralized Controller protocol interface",
        "Requirement": "Centralized Controller must be able to get interface with other PLC as per the future requirements via Ethernet or Modbus RS 485 protocol.",
        "상세 내용": "Centralized controllers must support standard Modbus RS485 and Ethernet communication protocols for interfacing with external PLCs."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "Technical Specifications: Part-A / Section M: Civil / 5 Buildings & Structures / Item (iii)",
        "Page#": "47",
        "Item": "Control Room Civil Design - Simhadri",
        "Requirement": "Electrical cum Control Room for Carbon capture plant at Simadhari: A RCC / pre-engineered structural steel Building (25m X 12m) tentative shall be provided. The building shall be covered from all sides. The Electrical Room shall accommodate, LT PMCC Panels, Dry Type Transformer (if applicable), VFD Transformers (if applicable), 220V DCFB, LDB, WDB. and UPS, UPS DCDB, UPS battery and operator cabin etc. as per functional requirement. However, final size of Switchgear room, Control room, individual room, partition, doors, windows, rolling shutter etc shall be finalized at the stage of detail engineering with the approval of NTPC. Control room, meeting room, office space with toilet, pantry, staircase etc., shall be provided with sufficient lighting, ceiling / wall mounted fans, office furniture (tables, chairs etc of reputed make) etc. Individual control room, meeting room, office space shall be able to house 5-8 people for regular operational requirement. Control Room shall be Air conditioned. The control shall accommodate all the equipment as mentioned in Electrical and C&I chapter. The firefighting system shall be provided as per norms.",
        "상세 내용": "Tentative 25x12m air-conditioned building at Simhadri to house C&I and electrical systems, including UPS systems, batteries, and the operator cabin."
    },
    {
        "ITB File Name": "Part_A_Technical_Specification_150_TPD_CO2_To_Urea.pdf",
        "Clause or Drawing No.": "Technical Specifications: Part-A / Section M: Civil / 5 Buildings & Structures / Item (iv)",
        "Page#": "47",
        "Item": "Control Room Civil Design - Pudimadaka",
        "Requirement": "Electrical cum Control Room for Urea Plant at Pudimadka shall be as mentioned in technical Specification (Part-B). Bidder shall be free to locate civil structures including control room in a manner to avoid blast proof constructions. However, compliance of OISD and other relevant statutory and mandatory requirements shall be ensured.",
        "상세 내용": "Pudimadaka Control Room civil design specification, with layout flexibility to avoid blast-proof construction while maintaining OISD compliance."
    }
]

# Save updated JSON database
print("Saving transformed JSON database...")
with open(output_json, "w", encoding="utf-8") as f:
    json.dump(new_records, f, indent=2, ensure_ascii=False)

# Save transformed CSVs
headers = ["ITB File Name", "Clause or Drawing No.", "Page#", "Item", "Requirement", "상세 내용"]

print("Saving CSV to brain...")
with open(output_csv_brain, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for rec in new_records:
        writer.writerow([rec[h] for h in headers])

print("Saving CSV to workspace...")
with open(output_csv_ws, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for rec in new_records:
        writer.writerow([rec[h] for h in headers])

# Save to Excel mimicking the example exactly
print("Loading example styling template...")
wb_ex = load_workbook(example_xlsx)
ws_ex = wb_ex.active

# Capture template styles
styles_header = []
for c in range(1, 7):
    cell = ws_ex.cell(1, c)
    styles_header.append({
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "alignment": copy.copy(cell.alignment),
        "border": copy.copy(cell.border)
    })

styles_row2 = []
for c in range(1, 7):
    cell = ws_ex.cell(2, c)
    styles_row2.append({
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "alignment": copy.copy(cell.alignment),
        "border": copy.copy(cell.border)
    })

styles_data = []
for c in range(1, 7):
    cell = ws_ex.cell(3, c) # Row 3 is a data row in the example
    styles_data.append({
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "alignment": copy.copy(cell.alignment),
        "border": copy.copy(cell.border)
    })

wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "Sheet1"

# Enable grid lines
ws_new.views.sheetView[0].showGridLines = True

# Copy row 1 height
ws_new.row_dimensions[1].height = ws_ex.row_dimensions[1].height
# Set column widths to 13.0 as in the example
for c in range(1, 7):
    col_letter = get_column_letter(c)
    ws_new.column_dimensions[col_letter].width = 13.0

# Write Row 1 (Headers)
ws_new.append(headers)
for c in range(1, 7):
    cell = ws_new.cell(1, c)
    cell.font = styles_header[c-1]["font"]
    cell.fill = styles_header[c-1]["fill"]
    cell.alignment = styles_header[c-1]["alignment"]
    cell.border = styles_header[c-1]["border"]

# Write Row 2 (Blank Spacer Row with borders)
ws_new.append([None]*6)
for c in range(1, 7):
    cell = ws_new.cell(2, c)
    cell.font = styles_row2[c-1]["font"]
    cell.fill = styles_row2[c-1]["fill"]
    cell.alignment = styles_row2[c-1]["alignment"]
    cell.border = styles_row2[c-1]["border"]

# Write Row 3 onwards (Data)
for idx, rec in enumerate(new_records, 3):
    # XML character scrubbing for safety
    row_data = []
    for h in headers:
        val = rec[h]
        if isinstance(val, str):
            val = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]', '', val)
        row_data.append(val)
        
    ws_new.append(row_data)
    
    # Apply styling
    for c in range(1, 7):
        cell = ws_new.cell(idx, c)
        cell.font = styles_data[c-1]["font"]
        cell.fill = styles_data[c-1]["fill"]
        cell.alignment = styles_data[c-1]["alignment"]
        cell.border = styles_data[c-1]["border"]

# Save to both workspace and brain paths
wb_new.save(output_xlsx_ws)
wb_new.save(output_xlsx_brain)
print("Excel files successfully saved to workspace and brain.")

# Save Markdown Table
print("Generating Markdown table...")
md_content = []
md_content.append("# Extracted Telecom Reference Data\n")
md_content.append("This table lists all telecom-related specifications, requirements, and clauses extracted from the available PDF references in the project directory, structured in accordance with the ITB summary example.\n")
md_content.append("## Summary Metrics\n")
md_content.append(f"- **Total Extracted Requirements**: {len(new_records)} clauses")

# Count by file
file_counts = {}
for rec in new_records:
    file_counts[rec["ITB File Name"]] = file_counts.get(rec["ITB File Name"], 0) + 1

sorted_files = sorted(file_counts.items(), key=lambda x: x[1], reverse=True)
for fname, count in sorted_files:
    md_content.append(f"  - {fname}: {count} clauses")

md_content.append("\n## Requirement Table\n")

# Build Table headers
md_headers = ["ITB File Name", "Clause or Drawing No.", "Page#", "Item", "Requirement", "상세 내용"]
md_content.append("| " + " | ".join(md_headers) + " |")
md_content.append("| " + " | ".join([":---" for _ in md_headers]) + " |")

# Build rows
for rec in new_records:
    # Format description to escape vertical bars and replace newlines
    req = rec["Requirement"].replace("|", "&#124;").replace("\n", " ").strip()
    note = rec["상세 내용"].replace("|", "&#124;").replace("\n", " ").strip()
    
    filename_url = f"file:///g:/My%20Drive/Project/CTGU/{rec['ITB File Name'].replace(' ', '%20')}"
    filename_link = f"[{rec['ITB File Name']}]({filename_url})"
    
    row_str = f"| {filename_link} | {rec['Clause or Drawing No.']} | {rec['Page#']} | {rec['Item']} | {req} | {note} |"
    md_content.append(row_str)

with open(output_md_brain, "w", encoding="utf-8") as f:
    f.write("\n".join(md_content) + "\n")
print("Markdown table successfully saved.")
