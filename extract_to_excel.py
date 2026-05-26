import os
import re
import json
import openpyxl
import copy
import sys
import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure Tesseract-OCR is in the PATH environment for Windows executions
tess_path = r"C:\Program Files\Tesseract-OCR"
if os.path.exists(tess_path) and tess_path not in os.environ["PATH"]:
    os.environ["PATH"] += os.pathsep + tess_path

# Import hifi_extractor modules
from hifi_extractor.page_classifier import classify_page, PageCategory
from hifi_extractor.text_extractor import extract_page
from hifi_extractor.pipeline import run_pipeline, PipelineConfig
import fitz

# Define directories and files
project_dir = os.path.dirname(os.path.abspath(__file__))
mapping_file = os.path.join(project_dir, "pdf_page_specs_mapping.json")
example_xlsx = os.path.join(project_dir, "ITB_SUMMARY_EXAMPLE.xlsx")
output_xlsx = os.path.join(project_dir, "telecom_extracted_requirements.xlsx")
db_file = os.path.join(project_dir, "telecom_extracted_requirements_db.json")

# Load mapping lookups
if os.path.exists(mapping_file):
    with open(mapping_file, "r", encoding="utf-8") as f:
        page_mappings = json.load(f)
else:
    page_mappings = {}

# Path to filter config
config_file = os.path.join(project_dir, "filter_config.json")

# Default hardcoded fallbacks
DEFAULT_KEYWORDS = [
    r"\bcctv\b", r"\bcamera[s]?\b", r"\bnvr\b", r"\bvms\b", r"\bsurveillance\b",
    r"\bpaga\b", r"\bpa/ga\b", r"\bpublic\s+address\b", r"\bgeneral\s+alarm\b", r"\bloudspeaker[s]?\b", r"\bacoustic\s+hood\b",
    r"\btelephone[s]?\b", r"\btelephony\b", r"\bintercom\b", r"\bpabx\b", r"\bhandset[s]?\b", r"\bhooter\b", r"\bflasher\b",
    r"\bfiber\s+optic[s]?\b", r"\bfoc\b", r"\boptical\s+fiber\b", r"\bstructured\s+cabling\b",
    r"\bcybersecurity\b", r"\bfirewall[s]?\b", r"\bnetwork\s+switch[es]?\b", r"\blan/wan\b", r"\bdmz\b", r"\biec\s+62443\b",
    r"\baccess\s+control\b", r"\bacs\b", r"\bcard\s+reader[s]?\b",
    r"\btelecom\b", r"\btelecommunication[s]?\b"
]

DEFAULT_FP_PATTERNS = [
    r"\bvalve material specification\b",
    r"\bpiping material specification\b",
    r"\bpms\s*&\s*vms\b",
    r"\bwelding specification\b",
    r"\bbolt tension\b",
    r"\bbolt torque\b",
    r"\bflange joint\b",
    r"\bmobile toilet\b",
    r"\beffluent summary\b",
    r"\bwater balance\b",
    r"\bbarc guideline\b",
    r"\bradiography source pit\b",
    r"\bfilm processing\b",
    r"\bclash checking\b",
    r"\b3-d modelling\b",
    r"\bhutments\b",
    r"\blabour colony\b",
    r"\bcanteen facility\b",
    r"\boffice facility\b"
]

DEFAULT_CATEGORY_ORDER = {
    "CCTV Surveillance System": 1,
    "Access Control System (ACS)": 2,
    "Public Address & General Alarm (PAGA) System": 3,
    "Telephone Intercom System": 4,
    "OT Network & Cybersecurity": 5,
    "Structured Cabling & FOC": 6,
    "UPS & DC Power Systems": 7,
    "Control Room Civil / Environmental": 8,
    "Telecom Specifications": 9
}

# Dynamic configurations loaded on start
KEYWORDS = list(DEFAULT_KEYWORDS)
FALSE_POSITIVE_PATTERNS = list(DEFAULT_FP_PATTERNS)
CATEGORY_ORDER = dict(DEFAULT_CATEGORY_ORDER)
BYPASS_FILTERING = False
NO_FILTER = False

def load_filter_config():
    global KEYWORDS, FALSE_POSITIVE_PATTERNS, CATEGORY_ORDER, BYPASS_FILTERING, NO_FILTER
    if os.path.exists(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                config_data = json.load(f)
            
            KEYWORDS = config_data.get("keywords", DEFAULT_KEYWORDS)
            FALSE_POSITIVE_PATTERNS = config_data.get("false_positive_patterns", DEFAULT_FP_PATTERNS)
            CATEGORY_ORDER = config_data.get("category_order", DEFAULT_CATEGORY_ORDER)
            BYPASS_FILTERING = config_data.get("bypass_filtering", False)
            NO_FILTER = config_data.get("no_filter", False)
            
            print("Successfully loaded filter configuration from filter_config.json.")
        except Exception as e:
            print(f"[WARNING] Error reading filter_config.json: {e}. Using defaults.")
            KEYWORDS = list(DEFAULT_KEYWORDS)
            FALSE_POSITIVE_PATTERNS = list(DEFAULT_FP_PATTERNS)
            CATEGORY_ORDER = dict(DEFAULT_CATEGORY_ORDER)
            BYPASS_FILTERING = False
            NO_FILTER = False
    else:
        print("filter_config.json not found. Using default configurations.")
        KEYWORDS = list(DEFAULT_KEYWORDS)
        FALSE_POSITIVE_PATTERNS = list(DEFAULT_FP_PATTERNS)
        CATEGORY_ORDER = dict(DEFAULT_CATEGORY_ORDER)
        BYPASS_FILTERING = False
        NO_FILTER = False

# Load config immediately on import
load_filter_config()

COMPILED_KWS = [re.compile(kw, re.IGNORECASE) for kw in KEYWORDS]
COMPILED_FP_PATS = [re.compile(pat, re.IGNORECASE) for pat in FALSE_POSITIVE_PATTERNS]

def reload_config_and_compile():
    """Allows on-demand reloading from configuration file (e.g. for Flask runs)"""
    global COMPILED_KWS, COMPILED_FP_PATS
    load_filter_config()
    COMPILED_KWS = [re.compile(kw, re.IGNORECASE) for kw in KEYWORDS]
    COMPILED_FP_PATS = [re.compile(pat, re.IGNORECASE) for pat in FALSE_POSITIVE_PATTERNS]

def is_telecom_clause(text):
    for pat in COMPILED_KWS:
        if pat.search(text):
            return True
    return False

# Anonymization replacement filter
def anonymize(text):
    if not isinstance(text, str):
        return text
    # Simhadri / Simadhari -> Site Alpha
    text = re.sub(r'Simhadri', 'Site Alpha', text, flags=re.IGNORECASE)
    text = re.sub(r'Simadhari', 'Site Alpha', text, flags=re.IGNORECASE)
    # Pudimadaka / Pudimadka -> Site Beta
    text = re.sub(r'Pudimadaka', 'Site Beta', text, flags=re.IGNORECASE)
    text = re.sub(r'Pudimadka', 'Site Beta', text, flags=re.IGNORECASE)
    # NTPC -> Client / Employer
    text = re.sub(r'\bNTPC\b', 'Client', text)
    # CTGU -> Project Alpha
    text = re.sub(r'\bCTGU\b', 'Project Alpha', text, flags=re.IGNORECASE)
    return text

# Group keyword classification for generating short Item labels
# Category order is now loadable dynamically but we keep CATEGORY_ORDER globally visible

def is_false_positive(text, bypass=None):
    if bypass is None:
        bypass = BYPASS_FILTERING
    if bypass:
        return False
        
    text_lower = text.lower()
    for pat in COMPILED_FP_PATS:
        if pat.search(text):
            return True
            
    if "vms" in text_lower:
        valve_context = ["valve", "piping", "fitting", "flange", "pms", "gasket", "material specification", "pressure rating", "asme"]
        if any(w in text_lower for w in valve_context):
            return True
            
    if "click on the document title" in text_lower or "table of contents" in text_lower:
        return True
        
    if "standard drawing list" in text_lower or "list of engineering" in text_lower or "standard specifications list" in text_lower:
        if len(re.findall(r'\bB773-\d+-\d+-\d+-\d+\b', text)) > 3 or len(re.findall(r'\bB773-\d+-\d+-\d+-SP-\d+\b', text)) > 3:
            return True
            
    return False

def get_item_label(text):
    text_lower = text.lower()
    if "cctv" in text_lower or "camera" in text_lower or "nvr" in text_lower:
        return "CCTV Surveillance System"
    if "paga" in text_lower or "loudspeaker" in text_lower or "public address" in text_lower:
        return "Public Address & General Alarm (PAGA) System"
    if "telephone" in text_lower or "telephony" in text_lower or "intercom" in text_lower or "pabx" in text_lower:
        return "Telephone Intercom System"
    if "fiber" in text_lower or "foc" in text_lower or "cabling" in text_lower or "cable" in text_lower:
        return "Structured Cabling & FOC"
    if "cybersecurity" in text_lower or "firewall" in text_lower or "switch" in text_lower or "dmz" in text_lower:
        return "OT Network & Cybersecurity"
    if "ups" in text_lower or "battery" in text_lower or "charger" in text_lower:
        return "UPS & DC Power Systems"
    if "access control" in text_lower or "acs" in text_lower or "card reader" in text_lower:
        return "Access Control System (ACS)"
    if "control room" in text_lower or "civil" in text_lower or "environmental" in text_lower:
        return "Control Room Civil / Environmental"
    return "Telecom Specifications"

def parse_page_number(page_str):
    if not page_str:
        return 999999
    match = re.search(r'\d+', str(page_str))
    if match:
        return int(match.group(0))
    return 999999

def sort_and_arrange_records(records):
    def get_sort_key(r):
        cat = r.get("Item", "Telecom Specifications")
        matched_cat = "Telecom Specifications"
        for key in CATEGORY_ORDER.keys():
            if key.lower() in cat.lower():
                matched_cat = key
                break
        cat_order = CATEGORY_ORDER.get(matched_cat, 9)
        file_name = r.get("ITB File Name", "")
        page_num = parse_page_number(r.get("Page#", ""))
        return (cat_order, file_name, page_num)
    
    return sorted(records, key=get_sort_key)


def generate_excel_table(records, template_path, output_path):
    """Generates the final Excel table copying styles from the template."""
    print("Loading style template from ITB_SUMMARY_EXAMPLE.xlsx...")
    if os.path.exists(template_path):
        wb_ex = load_workbook(template_path)
        ws_ex = wb_ex.active
    else:
        # Fallback empty workbook if template is not found
        wb_ex = Workbook()
        ws_ex = wb_ex.active
        # Write headers
        ws_ex.append(["ITB File Name", "Clause or\nDrawing No.", "Page#", "Item", "Requirement", "상세 내용"])
        # Format headers with default styling
        for col in range(1, 7):
            cell = ws_ex.cell(1, col)
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = PatternFill(fill_type="solid", start_color="D3D3D3", end_color="D3D3D3")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Read styling structures
    headers = ["ITB File Name", "Clause or\nDrawing No.", "Page#", "Item", "Requirement", "상세 내용"]
    
    styles_header = []
    for c in range(1, 7):
        cell = ws_ex.cell(1, c)
        styles_header.append({
            "font": copy.copy(cell.font),
            "fill": copy.copy(cell.fill),
            "alignment": copy.copy(cell.alignment),
            "border": copy.copy(cell.border)
        })

    styles_row2 = []
    for c in range(1, 7):
        cell = ws_ex.cell(2, c) if ws_ex.max_row >= 2 else ws_ex.cell(1, c)
        styles_row2.append({
            "font": copy.copy(cell.font),
            "fill": copy.copy(cell.fill),
            "alignment": copy.copy(cell.alignment),
            "border": copy.copy(cell.border)
        })

    styles_data = []
    for c in range(1, 7):
        cell = ws_ex.cell(3, c) if ws_ex.max_row >= 3 else ws_ex.cell(1, c)
        styles_data.append({
            "font": copy.copy(cell.font),
            "fill": copy.copy(cell.fill),
            "alignment": copy.copy(cell.alignment),
            "border": copy.copy(cell.border)
        })

    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Sheet1"
    
    # Enable Gridlines
    ws_new.views.sheetView[0].showGridLines = True
    
    # Set header row height and column widths
    ws_new.row_dimensions[1].height = ws_ex.row_dimensions[1].height if ws_ex.row_dimensions[1].height else 26.0
    for c in range(1, 7):
        col_letter = get_column_letter(c)
        ws_new.column_dimensions[col_letter].width = 13.0

    # Write Headers (Row 1)
    ws_new.append(headers)
    for c in range(1, 7):
        cell = ws_new.cell(1, c)
        cell.font = styles_header[c-1]["font"]
        cell.fill = styles_header[c-1]["fill"]
        cell.alignment = styles_header[c-1]["alignment"]
        cell.border = styles_header[c-1]["border"]

    # Write Blank Spacer (Row 2)
    ws_new.append([None]*6)
    for c in range(1, 7):
        cell = ws_new.cell(2, c)
        cell.font = styles_row2[c-1]["font"]
        cell.fill = styles_row2[c-1]["fill"]
        cell.alignment = styles_row2[c-1]["alignment"]
        cell.border = styles_row2[c-1]["border"]

    # Write Data (Row 3 onwards)
    for idx, rec in enumerate(records, 3):
        row_data = []
        for h in headers:
            val = rec[h]
            # Scrub invalid XML characters to prevent openpyxl write crash
            if isinstance(val, str):
                val = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]', '', val)
            row_data.append(val)
        
        ws_new.append(row_data)
        
        # Apply data cell styling
        for c in range(1, 7):
            cell = ws_new.cell(idx, c)
            cell.font = styles_data[c-1]["font"]
            cell.fill = styles_data[c-1]["fill"]
            cell.alignment = styles_data[c-1]["alignment"]
            cell.border = styles_data[c-1]["border"]

    wb_new.save(output_path)
    print(f"Excel table successfully saved to {output_path} with {len(records)} clauses.")
    
    # Save CSV version
    import csv
    output_csv = output_path.replace('.xlsx', '.csv')
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for rec in records:
            row_data = []
            for h in headers:
                val = rec[h]
                if isinstance(val, str):
                    val = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f-\x9f]', '', val)
                row_data.append(val)
            writer.writerow(row_data)
    print(f"CSV version successfully saved to {output_csv}.")

def run_extraction_suite(bypass_filtering=None, no_filter=None):
    if no_filter is None:
        no_filter = NO_FILTER
    if bypass_filtering is None:
        bypass_filtering = BYPASS_FILTERING

    pdf_files = sorted([f for f in os.listdir(project_dir) if f.lower().endswith('.pdf')])
    print(f"Found {len(pdf_files)} PDF documents in project directory.")
    
    all_records = []
    
    # Configure Pipeline Config (OCR and Margin suppression active)
    config = PipelineConfig(
        ocr_language="eng",
        suppress_margins=True,
        margin_method="static_zone",
        header_zone_pct=0.05,
        footer_zone_pct=0.05,
        format_markdown=True,
        enable_chunking=False
    )
    
    for filename in pdf_files:
        filepath = os.path.join(project_dir, filename)
        print(f"Processing document: {filename}...")
        
        doc = fitz.open(filepath)
        total_pages = len(doc)
        
        # Look up pre-calculated page specs if available
        file_mapping = page_mappings.get(filename, {})
        
        for p_idx in range(total_pages):
            page_num = p_idx + 1
            page = doc[p_idx]
            
            # Use page classification and extraction module
            classification = classify_page(page, page_num)
            
            # If the page classification is empty or contains negligible text, we skip to save time
            if classification.category == PageCategory.EMPTY:
                continue
                
            page_result = extract_page(doc, p_idx, classification=classification)
            text = page_result.text
            
            # Search for keyword matches in sentences/paragraphs
            paragraphs = re.split(r'\n\s*\n|\n(?=\d+\.)', text)
            for para in paragraphs:
                para = para.strip().replace('\n', ' ')
                para = re.sub(r'\s+', ' ', para)
                
                is_relevant = False
                if len(para) > 15:
                    if no_filter:
                        is_relevant = True
                    else:
                        is_relevant = is_telecom_clause(para) and not is_false_positive(para, bypass=bypass_filtering)
                
                if is_relevant:
                    # Get mapping lookup metadata
                    p_map = file_mapping.get(str(page_num), {})
                    doc_no = p_map.get("doc_no", "")
                    title = p_map.get("title", "")
                    
                    # Clause or drawing number formatting
                    clause_or_drg = ""
                    if doc_no and title:
                        clause_or_drg = f"{doc_no}, {title}"
                    elif doc_no:
                        clause_or_drg = doc_no
                    elif title:
                        clause_or_drg = title
                    else:
                        clause_or_drg = f"Specification Section / Page {page_num}"
                        
                    # Extract page reference info
                    # Check if there is printed page number, e.g. Page 4 of 20
                    printed_page_match = re.search(r'Page\s+(\d+)\s+of\s+(\d+)', text)
                    printed_page = f"Page {printed_page_match.group(1)}" if printed_page_match else f"{page_num}"
                    
                    # Crop/sanitize filename for sheet view (remove .pdf extension)
                    sheet_filename = filename.replace('.pdf', '')
                    
                    record = {
                        "ITB File Name": anonymize(sheet_filename),
                        "Clause or\nDrawing No.": anonymize(clause_or_drg),
                        "Page#": printed_page,
                        "Item": anonymize(get_item_label(para)),
                        "Requirement": anonymize(para),
                        "상세 내용": "Verbatim specification requirement extracted from ITB files."
                    }
                    all_records.append(record)
                    
        doc.close()
        
    sorted_records = sort_and_arrange_records(all_records)
    generate_excel_table(sorted_records, example_xlsx, output_xlsx)

def main():
    parser = argparse.ArgumentParser(description="Standalone Telecom Specification Extraction Tool")
    parser.add_argument("--force-extract", action="store_true", help="Force dynamic PDF extraction rather than using database cache")
    parser.add_argument("--bypass-filtering", action="store_true", help="Bypass all false-positive filtering")
    parser.add_argument("--no-filter", action="store_true", help="Do not filter any words; extract all paragraphs")
    args = parser.parse_args()

    # Default to loading pre-extracted database if it exists
    if not args.force_extract and os.path.exists(db_file):
        print(f"Loading pre-extracted database from {db_file}...")
        with open(db_file, "r", encoding="utf-8") as f:
            records = json.load(f)
        # Ensure all loaded values are anonymized and formatted correctly
        anonymized_records = []
        for r in records:
            clean_r = {}
            for k, v in r.items():
                if k == "Clause or Drawing No.":
                    # Rename key to include newline to match headers exactly
                    clean_r["Clause or\nDrawing No."] = anonymize(v)
                elif k == "Page#":
                    clean_r[k] = str(v)
                else:
                    clean_r[k] = anonymize(v)
            
            req_text = clean_r.get("Requirement", "")
            is_relevant = False
            if len(req_text) > 15:
                if args.no_filter:
                    is_relevant = True
                else:
                    is_relevant = is_telecom_clause(req_text) and not is_false_positive(req_text, bypass=args.bypass_filtering)
            
            if is_relevant:
                anonymized_records.append(clean_r)
        
        sorted_records = sort_and_arrange_records(anonymized_records)
        generate_excel_table(sorted_records, example_xlsx, output_xlsx)
    else:
        print("Starting dynamic offline extraction from PDF documents (this may take a few minutes)...")
        run_extraction_suite(bypass_filtering=args.bypass_filtering, no_filter=args.no_filter)

if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    main()
