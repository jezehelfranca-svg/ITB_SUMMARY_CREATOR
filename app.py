import os
import re
import json
import tempfile
import sys
import threading
import shutil
from pathlib import Path
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from google import genai
except ImportError:
    genai = None

# Ensure project root is in Python path
project_dir = os.path.dirname(os.path.abspath(__file__))
project_path = Path(project_dir).resolve()
if project_dir not in sys.path:
    sys.path.append(project_dir)

import extract_to_excel

GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")


def process_clause_with_gemini(record, client):
    req = record.get("Requirement", "")
    prompt = f"""
You are an expert telecom and instrumentation engineering assistant.
Analyze this technical specification requirement clause:
"{req}"

Please determine if this clause is actually relevant to plant telecommunications, plant security (CCTV, ACS), networking/cybersecurity, PAGA, telephone, or telecom power supply.
Many clauses are false positives (e.g. tables of contents, document indexes, piping/valves specifications matching 'VMS', generic construction management, mobile toilets, general safety/ITP procedures, general electrical switchgears/PMCC, etc.).

Produce a JSON object with the following keys:
1. "IsRelevant": true or false. Set to false if this is a false positive (not a real telecom/security system requirement).
2. "Category": One of the following categories that best fits this requirement:
   - "CCTV Surveillance System"
   - "Access Control System (ACS)"
   - "Public Address & General Alarm (PAGA) System"
   - "Telephone Intercom System"
   - "OT Network & Cybersecurity"
   - "Structured Cabling & FOC"
   - "UPS & DC Power Systems"
   - "Control Room Civil / Environmental"
   - "Telecom Specifications"
3. "Item": A refined, specific "Item" label (max 5 words) that categorizes this telecom/security system or component (e.g. "CCTV Ingress Protection", "UPS Redundant Battery Backup", "PAGA Master Control Unit"). Do not use generic labels if a specific one is mentioned.
4. "상세 내용": A professional, technical summary in Korean for the summary column (focusing on key parameters, values, standards, and strictly replacing project/site names like "Simhadri", "Pudimadaka", or "CTGU" with "Site Alpha" or "Site Beta").

Format your response exactly as a JSON object. Return ONLY the JSON object. Do not include markdown code block formatting (like ```json).
"""
    try:
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config={"response_mime_type": "application/json"},
        )
        text = response.text.strip()
        text = re.sub(r'^```json\s*|\s*```$', '', text, flags=re.IGNORECASE)
        data = json.loads(text)
        return {
            "IsRelevant": data.get("IsRelevant", True),
            "Category": data.get("Category", record.get("Item", "Telecom Specifications")),
            "Item": data.get("Item", record.get("Item", "")),
            "상세 내용": data.get("상세 내용", record.get("상세 내용", ""))
        }
    except Exception as e:
        print(f"Gemini API error on clause: {e}")
        return {
            "IsRelevant": True,
            "Category": record.get("Item", "Telecom Specifications"),
            "Item": record.get("Item", ""),
            "상세 내용": record.get("상세 내용", "")
        }

def polish_clauses_with_ai(records, api_key):
    if genai is None:
        raise RuntimeError(
            "AI polishing requires the optional google-genai dependency."
        )

    print("Configuring Gemini API key and launching AI polisher...")
    client = genai.Client(api_key=api_key)
    
    polished_records = [dict(record) for record in records]
    irrelevant_indices = set()
    
    def process_index(idx, rec):
        print(f"AI polishing clause {idx+1}/{len(records)}...")
        try:
            res = process_clause_with_gemini(rec, client)
            if not res["IsRelevant"]:
                print(f"AI marked clause {idx+1} as IRRELEVANT. Filtering out.")
                return idx, None, False
            new_rec = dict(rec)
            new_rec["Category"] = res["Category"]
            new_rec["Item"] = res["Item"]
            new_rec["상세 내용"] = res["상세 내용"]
            return idx, new_rec, True
        except Exception as e:
            print(f"Error in process_index for clause {idx+1}: {e}")
            return idx, rec, True

    try:
        with ThreadPoolExecutor(max_workers=8) as executor:
            futures = [executor.submit(process_index, i, r) for i, r in enumerate(records)]
            for fut in as_completed(futures):
                try:
                    idx, polished_rec, is_relevant = fut.result()
                    if is_relevant:
                        polished_records[idx] = polished_rec
                    else:
                        irrelevant_indices.add(idx)
                except Exception as e:
                    print(f"Error in AI worker thread: {e}. Keeping original clause.")
    finally:
        client.close()

    final_records = [
        record
        for idx, record in enumerate(polished_records)
        if idx not in irrelevant_indices
    ]
            
    print(f"AI polishing completed. {len(final_records)}/{len(records)} clauses retained.")
    return final_records

app = Flask(__name__, template_folder=os.path.join(project_dir, 'templates'))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB limit

# Global log buffer for UI streaming
ui_logs = []
logs_lock = threading.Lock()
status_lock = threading.Lock()
extraction_lock = threading.Lock()
config_lock = threading.Lock()
MAX_UI_LOGS = 2000
MAX_FILTER_PATTERNS = 500
MAX_FILTER_PATTERN_LENGTH = 500

# Initial output names based on current folder contents
pdf_files_init = sorted([f for f in os.listdir(project_dir) if f.lower().endswith('.pdf')])
default_base = pdf_files_init[0].replace('.pdf', '') if pdf_files_init else "telecom_extracted_requirements"

# Execution status dictionary to prevent frontend race conditions
execution_status = {
    "status": "idle",
    "count": 0,
    "error": None,
    "output_xlsx": f"{default_base}.xlsx",
    "output_csv": f"{default_base}.csv"
}

def append_ui_log(message):
    """Append a bounded log message for the UI."""
    clean_msg = str(message).strip()
    if not clean_msg:
        return
    with logs_lock:
        ui_logs.append(clean_msg)
        if len(ui_logs) > MAX_UI_LOGS:
            del ui_logs[:-MAX_UI_LOGS]


def update_execution_status(**updates):
    """Update shared extraction status atomically."""
    with status_lock:
        execution_status.update(updates)


def get_execution_status():
    """Return a consistent snapshot of shared extraction status."""
    with status_lock:
        return dict(execution_status)


def resolve_project_file(relative_path, required_suffix=None):
    """Resolve a user-provided relative path without leaving the project."""
    if not relative_path:
        raise ValueError("A file path is required.")

    candidate = (project_path / relative_path).resolve()
    try:
        candidate.relative_to(project_path)
    except ValueError as exc:
        raise ValueError("Invalid file path.") from exc

    if required_suffix and candidate.suffix.lower() != required_suffix.lower():
        raise ValueError(f"File must have {required_suffix} extension.")
    return candidate


def validate_filter_patterns(patterns, field_name, require_one=False):
    """Normalize and validate a user-edited list of regular expressions."""
    if not isinstance(patterns, list):
        raise ValueError(f"{field_name} must be a list.")

    normalized = []
    seen = set()
    for index, pattern in enumerate(patterns, 1):
        if not isinstance(pattern, str):
            raise ValueError(f"{field_name} entry {index} must be text.")

        pattern = pattern.strip()
        if not pattern or pattern in seen:
            continue
        if len(pattern) > MAX_FILTER_PATTERN_LENGTH:
            raise ValueError(
                f"{field_name} entry {index} exceeds "
                f"{MAX_FILTER_PATTERN_LENGTH} characters."
            )
        try:
            re.compile(pattern, re.IGNORECASE)
        except re.error as exc:
            raise ValueError(
                f"Invalid regex in {field_name} entry {index}: {exc}"
            ) from exc

        normalized.append(pattern)
        seen.add(pattern)

    if require_one and not normalized:
        raise ValueError("At least one inclusion keyword is required.")
    if len(normalized) > MAX_FILTER_PATTERNS:
        raise ValueError(
            f"{field_name} cannot contain more than {MAX_FILTER_PATTERNS} entries."
        )
    return normalized


def read_filter_config():
    """Read the persisted filter config, falling back to built-in defaults."""
    config_path = Path(extract_to_excel.config_file)
    config = {
        "bypass_filtering": False,
        "no_filter": False,
        "keywords": list(extract_to_excel.DEFAULT_KEYWORDS),
        "false_positive_patterns": list(extract_to_excel.DEFAULT_FP_PATTERNS),
        "category_order": dict(extract_to_excel.DEFAULT_CATEGORY_ORDER),
    }
    if config_path.is_file():
        with config_path.open("r", encoding="utf-8") as handle:
            saved_config = json.load(handle)
        if isinstance(saved_config, dict):
            config.update(saved_config)
    return config


def write_filter_config(config):
    """Atomically persist filter configuration."""
    config_path = Path(extract_to_excel.config_file)
    temp_path = config_path.with_suffix(".json.tmp")
    try:
        with temp_path.open("w", encoding="utf-8") as handle:
            json.dump(config, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
        temp_path.replace(config_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


class UILogRedirector:
    def __init__(self, original_stdout):
        self.original_stdout = original_stdout

    def write(self, message):
        self.original_stdout.write(message)
        append_ui_log(message)

    def flush(self):
        self.original_stdout.flush()

# Redirect stdout to capture logs for the web UI
sys.stdout = UILogRedirector(sys.stdout)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/list-pdfs', methods=['GET'])
def list_pdfs():
    try:
        files = sorted([f for f in os.listdir(project_dir) if f.lower().endswith('.pdf')])
        # Add basic info about whether they are in mapping
        mapping = extract_to_excel.page_mappings
        pdf_list = []
        for f in files:
            size_mb = os.path.getsize(os.path.join(project_dir, f)) / (1024 * 1024)
            has_mapping = f in mapping
            pdf_list.append({
                "name": f,
                "size_mb": round(size_mb, 2),
                "has_mapping": has_mapping
            })
        return jsonify({"success": True, "pdfs": pdf_list})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/logs', methods=['GET'])
def get_logs():
    clear = request.args.get('clear', 'false').lower() == 'true'
    with logs_lock:
        logs = list(ui_logs)
        if clear:
            ui_logs.clear()
    return jsonify({"success": True, "logs": logs})

@app.route('/api/status', methods=['GET'])
def get_status():
    return jsonify(get_execution_status())

@app.route('/api/filter-config', methods=['GET'])
def get_filter_config():
    try:
        with config_lock:
            config = read_filter_config()
        return jsonify({
            "success": True,
            "keywords": config["keywords"],
            "false_positive_patterns": config["false_positive_patterns"],
            "defaults": {
                "keywords": extract_to_excel.DEFAULT_KEYWORDS,
                "false_positive_patterns": extract_to_excel.DEFAULT_FP_PATTERNS,
            },
        })
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route('/api/filter-config', methods=['PUT'])
def update_filter_config():
    if not extraction_lock.acquire(blocking=False):
        return jsonify({
            "success": False,
            "error": "Filter rules cannot be changed while extraction is running."
        }), 409

    try:
        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            return jsonify({
                "success": False,
                "error": "A JSON filter configuration is required."
            }), 400

        keywords = validate_filter_patterns(
            payload.get("keywords"),
            "inclusion keywords",
            require_one=True,
        )
        false_positive_patterns = validate_filter_patterns(
            payload.get("false_positive_patterns"),
            "exclusion patterns",
        )

        with config_lock:
            config = read_filter_config()
            config["keywords"] = keywords
            config["false_positive_patterns"] = false_positive_patterns
            write_filter_config(config)
            extract_to_excel.reload_config_and_compile()

        append_ui_log(
            f"Filter rules updated: {len(keywords)} inclusion keywords and "
            f"{len(false_positive_patterns)} exclusion patterns."
        )
        return jsonify({
            "success": True,
            "message": "Filter rules saved and activated.",
            "keyword_count": len(keywords),
            "exclusion_count": len(false_positive_patterns),
        })
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    finally:
        extraction_lock.release()


@app.route('/api/extract', methods=['POST'])
def start_extraction():
    selected_files = list(dict.fromkeys(request.form.getlist('files')))
    force_extract = request.form.get('force_extract', 'false').lower() == 'true'
    enable_ai = request.form.get('enable_ai', 'false').lower() == 'true'
    api_key = request.form.get('api_key', '').strip()
    bypass_filtering = request.form.get('bypass_filtering', 'false').lower() == 'true'
    no_filter = request.form.get('no_filter', 'false').lower() == 'true'
    uploaded_file = request.files.get('file')

    local_pdfs = {
        path.name: path
        for path in project_path.iterdir()
        if path.is_file() and path.suffix.lower() == '.pdf'
    }
    invalid_files = [name for name in selected_files if name not in local_pdfs]
    if invalid_files:
        return jsonify({
            "success": False,
            "error": f"Invalid local PDF selection: {invalid_files[0]}"
        }), 400

    upload_name = ""
    if uploaded_file and uploaded_file.filename:
        upload_name = secure_filename(uploaded_file.filename)
        if not upload_name or Path(upload_name).suffix.lower() != '.pdf':
            return jsonify({
                "success": False,
                "error": "Uploaded file must be a PDF."
            }), 400

    if not selected_files and not upload_name:
        return jsonify({
            "success": False,
            "error": "Select or upload at least one PDF."
        }), 400

    if enable_ai and not api_key:
        return jsonify({
            "success": False,
            "error": "Enter a Gemini API key before enabling AI polishing."
        }), 400

    if enable_ai and genai is None:
        return jsonify({
            "success": False,
            "error": "Install google-genai to use AI polishing."
        }), 503

    if not extraction_lock.acquire(blocking=False):
        return jsonify({
            "success": False,
            "error": "An extraction job is already running."
        }), 409

    upload_dir = None
    try:
        # Reload dynamic filtering config only after this job owns the worker
        # lock so another request cannot mutate active extraction rules.
        with config_lock:
            extract_to_excel.reload_config_and_compile()

        targets = [
            {"name": name, "path": str(local_pdfs[name]), "use_cache": True}
            for name in selected_files
        ]

        # Save uploads while the request stream is still open. The worker only
        # receives a stable filesystem path and removes it after processing.
        if upload_name:
            upload_dir = tempfile.mkdtemp(prefix=".itb_upload_", dir=project_dir)
            upload_path = Path(upload_dir) / upload_name
            uploaded_file.save(upload_path)
            targets.append({
                "name": upload_name,
                "path": str(upload_path),
                "use_cache": False
            })

        first_fname = upload_name or selected_files[0]
        base_name = secure_filename(Path(first_fname).stem) or "telecom_extracted_requirements"
        output_xlsx = f"{base_name}.xlsx"
        output_csv = f"{base_name}.csv"

        update_execution_status(
            status="running",
            count=0,
            error=None,
            output_xlsx=output_xlsx,
            output_csv=output_csv,
        )

        with logs_lock:
            ui_logs.clear()
        append_ui_log("Initializing extraction runner...")
    except Exception as exc:
        extraction_lock.release()
        if upload_dir:
            shutil.rmtree(upload_dir, ignore_errors=True)
        return jsonify({"success": False, "error": str(exc)}), 500

    def run_async_extraction():
        try:
            print(f"Target files to extract: {', '.join(t['name'] for t in targets)}")

            # Load the reference database if exists
            db_records = []
            if os.path.exists(extract_to_excel.db_file):
                with open(extract_to_excel.db_file, "r", encoding="utf-8") as f:
                    db_records = json.load(f)

            all_records = []
            
            for target in targets:
                fname = target["name"]
                filepath = target["path"]
                sheet_filename = Path(fname).stem
                records_before_file = len(all_records)

                # Uploaded documents always run through extraction so a stale
                # cache entry with the same filename cannot replace their data.
                cached_items = [r for r in db_records if r.get('ITB File Name', '').replace('.pdf', '') == sheet_filename]
                
                if not force_extract and target["use_cache"] and cached_items:
                    print(f"Found {len(cached_items)} cached records in database for {fname}. Loading instantly...")
                    filtered_cached = [r for r in cached_items if no_filter or not extract_to_excel.is_false_positive(r.get("Requirement", ""), bypass=bypass_filtering)]
                    all_records.extend(filtered_cached)
                else:
                    # Run dynamic page-by-page parser
                    if not os.path.exists(filepath):
                        print(f"Error: PDF file not found at {filepath}")
                        continue
                    
                    import fitz
                    from hifi_extractor.page_classifier import classify_page, PageCategory
                    from hifi_extractor.text_extractor import extract_page
                    
                    doc = fitz.open(filepath)
                    try:
                        total_pages = len(doc)
                        file_mapping = extract_to_excel.page_mappings.get(fname, {})

                        print(f"Scanning {total_pages} pages of {fname} dynamically...")

                        for p_idx in range(total_pages):
                            page_num = p_idx + 1
                            page = doc[p_idx]

                            classification = classify_page(page, page_num)
                            if classification.category == PageCategory.EMPTY:
                                continue

                            # Report progress
                            if page_num % 10 == 0 or page_num == total_pages:
                                print(f"[{fname}] Processing page {page_num}/{total_pages}...")

                            page_result = extract_page(doc, p_idx, classification=classification)
                            text = page_result.text

                            paragraphs = re.split(r'\n\s*\n|\n(?=\d+\.)', text)
                            for para in paragraphs:
                                para = para.strip().replace('\n', ' ')
                                para = re.sub(r'\s+', ' ', para)

                                is_relevant = False
                                if len(para) > 15:
                                    if no_filter:
                                        is_relevant = True
                                    else:
                                        is_relevant = extract_to_excel.is_telecom_clause(para) and not extract_to_excel.is_false_positive(para, bypass=bypass_filtering)

                                if is_relevant:
                                    p_map = file_mapping.get(str(page_num), {})
                                    doc_no = p_map.get("doc_no", "")
                                    title = p_map.get("title", "")

                                    clause_or_drg = ""
                                    if doc_no and title:
                                        clause_or_drg = f"{doc_no}, {title}"
                                    elif doc_no:
                                        clause_or_drg = doc_no
                                    elif title:
                                        clause_or_drg = title
                                    else:
                                        clause_or_drg = f"Specification Section / Page {page_num}"

                                    printed_page_match = re.search(r'Page\s+(\d+)\s+of\s+(\d+)', text)
                                    printed_page = f"Page {printed_page_match.group(1)}" if printed_page_match else f"{page_num}"

                                    record = {
                                        "ITB File Name": Path(fname).stem,
                                        "Clause or\nDrawing No.": clause_or_drg,
                                        "Page#": printed_page,
                                        "Item": extract_to_excel.get_item_label(para),
                                        "Requirement": para,
                                        "상세 내용": "Verbatim specification requirement extracted dynamically."
                                    }
                                    all_records.append(record)
                    finally:
                        doc.close()
                    found_count = len(all_records) - records_before_file
                    print(f"Dynamic extraction finished for {fname}. Found {found_count} clauses.")
 
            if enable_ai and api_key:
                try:
                    all_records = polish_clauses_with_ai(all_records, api_key)
                except Exception as ai_err:
                    print(f"AI polishing failed: {ai_err}. Falling back to default rules.")
 
            # Fully anonymize all collected records
            anonymized_records = []
            for r in all_records:
                clean_r = {}
                for k, v in r.items():
                    # Align key names to excel headers
                    excel_key = "Clause or\nDrawing No." if k in ["Clause or Drawing No.", "Clause or\nDrawing No."] else k
                    clean_r[excel_key] = extract_to_excel.anonymize(str(v))
                anonymized_records.append(clean_r)
 
            # Sort the final list of anonymized records logically
            sorted_records = extract_to_excel.sort_and_arrange_records(anonymized_records)
 
            # Generate final Excel and CSV output files
            print("Compiling final styled spreadsheet...")
            output_xlsx_path = os.path.join(project_dir, output_xlsx)
            extract_to_excel.generate_excel_table(sorted_records, extract_to_excel.example_xlsx, output_xlsx_path)
            print("Finished writing output tables. Complete!")
            update_execution_status(status="completed", count=len(sorted_records))
            
        except Exception as e:
            print(f"Error in extraction process: {e}")
            update_execution_status(status="failed", error=str(e))
        finally:
            if upload_dir:
                shutil.rmtree(upload_dir, ignore_errors=True)
            extraction_lock.release()

    # Launch extraction thread to avoid blocking Flask
    t = threading.Thread(target=run_async_extraction, daemon=True)
    t.start()
    return jsonify({"success": True, "message": "Extraction started in background."})

@app.route('/api/preview-data', methods=['GET'])
def preview_data():
    try:
        status = get_execution_status()
        output_xlsx_path = os.path.join(project_dir, status.get("output_xlsx", "telecom_extracted_requirements.xlsx"))
        if not os.path.exists(output_xlsx_path):
            return jsonify({"success": False, "error": "No output spreadsheet generated yet."})
        
        # Read the generated spreadsheet rows
        wb = load_workbook(output_xlsx_path, data_only=True)
        ws = wb.active
        rows = []
        headers = ["ITB File Name", "Clause or\nDrawing No.", "Page#", "Item", "Requirement", "상세 내용"]
        
        # Row 1 is headers, Row 2 is blank spacer, Row 3+ is data
        for r_idx in range(3, ws.max_row + 1):
            row_vals = [ws.cell(r_idx, c_idx).value for c_idx in range(1, 7)]
            if any(row_vals): # Skip fully empty rows
                rows.append({
                    "ITB File Name": row_vals[0] or "",
                    "Clause or Drawing No.": row_vals[1] or "",
                    "Page#": row_vals[2] or "",
                    "Item": row_vals[3] or "",
                    "Requirement": row_vals[4] or "",
                    "상세 내용": row_vals[5] or ""
                })
        return jsonify({"success": True, "rows": rows, "count": len(rows)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/download', methods=['GET'])
def download_file():
    file_format = request.args.get('format', 'xlsx').lower()
    status = get_execution_status()
    output_xlsx_name = status.get("output_xlsx", "telecom_extracted_requirements.xlsx")
    if file_format == 'csv':
        filepath = os.path.join(project_dir, output_xlsx_name.replace('.xlsx', '.csv'))
        mimetype = 'text/csv'
        download_name = output_xlsx_name.replace('.xlsx', '.csv')
    else:
        filepath = os.path.join(project_dir, output_xlsx_name)
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        download_name = output_xlsx_name

    if not os.path.exists(filepath):
        return jsonify({"success": False, "error": "Requested file does not exist. Run extraction first."}), 404
        
    return send_file(filepath, mimetype=mimetype, as_attachment=True, download_name=download_name)

@app.route('/diagram-maker')
def diagram_maker():
    return render_template('diagram_maker.html')

@app.route('/api/diagrams', methods=['GET'])
def list_diagrams():
    try:
        diagrams = []
        # Scan root folder
        for f in os.listdir(project_dir):
            if f.lower().endswith('.mmd'):
                diagrams.append({
                    "name": f,
                    "path": f,
                    "location": "root"
                })
        # Scan subfolder CTGU-main/CTGU-main
        subfolder = os.path.join(project_dir, "CTGU-main", "CTGU-main")
        if os.path.exists(subfolder):
            for f in os.listdir(subfolder):
                if f.lower().endswith('.mmd'):
                    diagrams.append({
                        "name": f,
                        "path": os.path.join("CTGU-main", "CTGU-main", f),
                        "location": "subfolder"
                    })
        return jsonify({"success": True, "diagrams": diagrams})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/diagrams/load', methods=['GET'])
def load_diagram():
    rel_path = request.args.get('path', '').strip()
    try:
        full_path = resolve_project_file(rel_path, '.mmd')
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    if not full_path.is_file():
        return jsonify({"success": False, "error": "File not found."}), 404
        
    try:
        with full_path.open("r", encoding="utf-8") as f:
            code = f.read()
        return jsonify({"success": True, "code": code})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/diagrams/save', methods=['POST'])
def save_diagram():
    rel_path = request.form.get('path', '').strip()
    code = request.form.get('code', '')
    
    try:
        full_path = resolve_project_file(rel_path, '.mmd')
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
        
    try:
        full_path.parent.mkdir(parents=True, exist_ok=True)
        with full_path.open("w", encoding="utf-8") as f:
            f.write(code)
        print(f"Diagram saved successfully to {full_path}")
        return jsonify({"success": True, "message": "Diagram saved successfully."})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/diagrams/export-pdf', methods=['POST'])
def export_diagram_pdf():
    import base64
    import zlib
    import requests
    
    rel_path = request.form.get('path', '').strip()
    code = request.form.get('code', '')
    
    try:
        full_path = resolve_project_file(rel_path, '.mmd')
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    pdf_path = full_path.with_suffix('.pdf')
    
    try:
        # Clean code to reduce URL size
        cleaned_lines = []
        for line in code.splitlines():
            stripped = line.strip()
            if stripped.startswith("%%") and not stripped.startswith("%%{"):
                continue
            if not stripped:
                continue
            cleaned_lines.append(stripped)
        cleaned_code = "\n".join(cleaned_lines)
        
        # Prepare state JSON
        state = {
            "code": cleaned_code,
            "mermaid": {"theme": "default"}
        }
        json_str = json.dumps(state, separators=(',', ':'))
        
        # Compress and base64 urlsafe encode
        compressed = zlib.compress(json_str.encode('utf-8'), level=9)
        encoded = base64.b64encode(compressed).decode('utf-8')
        encoded_urlsafe = encoded.replace('+', '-').replace('/', '_').replace('=', '')
        
        # Call mermaid.ink
        url = f"https://mermaid.ink/pdf/pako:{encoded_urlsafe}?fit&landscape&paper=a3"
        print(f"Requesting A3 PDF export from: {url[:100]}...")
        
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=60)
        
        if r.status_code == 200:
            with pdf_path.open("wb") as f:
                f.write(r.content)
            print(f"Exported PDF saved to {pdf_path}")
            return send_file(pdf_path, mimetype='application/pdf', as_attachment=True, download_name=pdf_path.name)
        else:
            return jsonify({
                "success": False, 
                "error": f"Failed to generate PDF from mermaid.ink. Status code: {r.status_code}",
                "detail": r.text[:200]
            }), 500
            
    except Exception as e:
        print(f"Error exporting PDF: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.errorhandler(413)
def upload_too_large(_error):
    return jsonify({
        "success": False,
        "error": "Uploaded PDF exceeds the 50 MB limit."
    }), 413

if __name__ == '__main__':
    print("Starting Flask web backend at http://localhost:5000...")
    app.run(host='127.0.0.1', port=5000, debug=False)
